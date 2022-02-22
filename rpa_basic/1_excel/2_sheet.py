from openpyxl import Workbook
wb = Workbook()
#wb.active
ws = wb.create_sheet() #새로운 시트 기본이름으로 생성
ws.title = "Mysheet"
ws.sheet_properties.tabColor = "66ffff"

ws1 = wb.create_sheet("YourSheet") #주어진 이름으로 시트 생성
ws2 = wb.create_sheet("NewSheet", 2)


new_ws  = wb["NewSheet"]

print(wb.sheetnames)

#sheet 복사
new_ws["A1"]  = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")