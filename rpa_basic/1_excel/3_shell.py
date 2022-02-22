from operator import index
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4

print(ws["A1"])  #A1 셀의 정보를 출력
print(ws["A1"].value)
print(ws["A10"].value)  #값이 없을 때는 'None' 출력

#row = 1, 2, 3...
#column = A(1), B(2), C(3) ...
print(ws.cell( column= 1,row= 1).value) #ws "A1". value와 같음
print(ws.cell( column= 2,row= 1).value) #ws "B1". value

c =ws.cell(column=3, row = 1,value= 10) #ws["C1"]
print(c.value)   #c1.value

from random import *

#반복문 이용해서 랜덤 숫자 채우기
for x in range(1, 11):#10개 row
    for y in range(1, 11): #10개 column
        ws.cell(row = x, column= y, value= index)
        index += 1

wb.save("sample.xlsx")