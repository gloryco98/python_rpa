from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i,randint(0, 100), randint(0, 100)])

col_B = ws["B"]   #영어 컬럼만 가지고 오기
#print(col_B)
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"] #영어 수학 컬럼 함께 가져오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)
row_title = ws[1]
for cell in row_title:
    print(cell.value)

row_range = ws[2:6]
# for rows in row_range:
#     for cell in row_title:
#         print(cell.value, end= "")
#         print()

row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        print(cell.value, end = " ")
        print()


wb.save("sample.xlsx")